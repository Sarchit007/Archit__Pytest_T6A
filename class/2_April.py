# import requests
# import pytest
#
# @pytest.mark.skip
# def test_tcheckPet():
#     u = "https://petstore.swagger.io/v2/pet/1"
#     resp = requests.get(u)
#     print(resp.status_code)
#     print("pet is avialable ")
#
#
# @pytest.mark.skipif(True, reason="Condition is False, so test will run")
# def test_addPet():
#     u = "https://petstore.swagger.io/v2/pet"
#     payload = {
#         "id": 2,
#         "category": {
#             "id": 2,
#             "name": "string"
#         },
#         "name": "rocky",
#         "photoUrls": [
#             "string"
#         ],
#         "tags": [
#             {
#                 "id": 2,
#                 "name": "string"
#             }
#         ],
#         "status": "available"
#     }
#     res = requests.post(u, json=payload)
#     assert res.status_code == 200
#     assert res.json()["id"] == payload["id"]
#     assert res.json()["name"] == "rocky"
#
# @pytest.mark.parametrize('a,b',[(1,2)])
# def test_add(a,b):
#     return a+b


import pytest
import openpyxl

wb = openpyxl.Workbook()
sheetName = "sheet1"
if sheetName in wb.sheetnames:
    ws = wb[sheetName]
else:
    ws = wb.create_sheet(sheetName)

ws['A1']   = "user"
ws['B1'] = "password"
ws.append(['user1','123'])
ws.append(['standard_user','secret_sauce'])

print(ws.cell(row=2,column=2).value)

for r in ws.iter_rows(values_only=True):
    print(r)

def get_test_data():
    wb = openpyxl.load_workbook("sample.xlsx")
    ws = wb["sheet1"]

    data = []

    for r in ws.iter_rows(max_row=2,values_only=True):
        data.append(r)


    return data